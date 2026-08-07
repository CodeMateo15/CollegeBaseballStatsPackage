"""Derive baseball program-finance features from the federal EADA survey.

Run as a script; importing this module has no side effects.

    python -m ncaa_bbStats.program_store --eada-dir "path/to/EADA Data"

Source: the Equity in Athletics Disclosure Act survey, U.S. Department of
Education, https://ope.ed.gov/athletics/. Every co-educational institution
receiving Title IV funding must file annually, so this is public-domain federal
data with near-complete coverage.

The workbooks are about 100 MB and 4,275 columns each. They are **not** shipped;
this reads them from wherever you downloaded them and writes only the derived
per-team-season features. Point ``--eada-dir`` at a directory containing
``EADA_<YYYY>.xlsx`` files (at any depth).

Year alignment: ``EADA_<YYYY>.xlsx`` covers the academic year ending in YYYY,
which is the spring YYYY baseball season -- so it maps onto the season year with
no lag. Institutions file the 2025-26 survey in October 2026, so the 2026 season
carries 2025 forward, flagged in every row.

Requires openpyxl, which comes with the ``scrape`` extra.
"""

import argparse
import csv
import math
import os
import re

from ncaa_bbStats._paths import data_path
from ncaa_bbStats.team_registry import resolve_team, team_info

REPORTING_YEARS = [2021, 2022, 2023, 2024, 2025]

# Seasons with no survey of their own, and which reporting year they reuse.
# Delete an entry once the real workbook lands.
CARRY_FORWARD = {2026: 2025}

SOURCE_COLUMNS = [
    "unitid", "institution_name", "state_cd", "classification_name",
    "PARTIC_MEN_Baseball", "OPEXPPERPART_MEN_Baseball", "EXP_MEN_Baseball",
    "REV_MEN_Baseball", "MEN_TOTAL_HEADCOACH_Baseball",
    "MEN_TOTAL_ASSTCOACH_Baseball", "RECRUITEXP_MEN", "HDCOACH_SAL_FTE_MEN",
]

FEATURE_COLUMNS = [
    "budget_pct", "log_budget", "opex_per_player_pct", "log_opex_per_player",
    "log_budget_per_player", "roster_size", "log_revenue", "net_revenue",
    "coaching_staff_size", "dept_recruiting_pct", "log_dept_recruiting",
    "log_dept_coach_salary",
]

OUTPUT_COLUMNS = (
    ["team_id", "institution_name", "unitid", "season", "eada_year",
     "carried_forward", "state"] + FEATURE_COLUMNS
)

# A system-wide filing sums every branch campus and reports an impossible
# baseball roster -- Penn State files 453 participants in 2021. No NCAA roster
# runs past about 60, so anything outside this band is treated as unreported
# rather than allowed to poison the per-player ratios.
PLAUSIBLE_ROSTER = (15, 75)


def _log1p_nonneg(value):
    """log(1 + x) with negatives floored at zero, or None."""
    if value is None:
        return None
    return round(math.log1p(max(0.0, value)), 6)


def _number(value):
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return None if math.isnan(result) else result


def _percentile_ranks(values):
    """Percentile rank within a list, ignoring None. Returns a parallel list.

    Ranks are taken within a reporting year rather than using raw dollars.
    Baseball budgets inflate a few percent a year, so raw figures would make the
    carried-forward season look systematically poorer than the others; the rank
    is stable across years and immune to that.
    """
    present = [(i, v) for i, v in enumerate(values) if v is not None]
    ranks = [None] * len(values)
    if not present:
        return ranks
    present.sort(key=lambda pair: pair[1])
    total = len(present)
    for position, (index, _value) in enumerate(present, start=1):
        ranks[index] = round(position / total, 6)
    return ranks


def find_workbooks(eada_dir: str) -> dict:
    """Locate EADA_<YYYY>.xlsx files under a directory, at any depth."""
    found = {}
    for root, _dirs, files in os.walk(eada_dir):
        for name in files:
            match = re.fullmatch(r"EADA_(\d{4})\.xlsx", name, re.IGNORECASE)
            if match:
                found[int(match.group(1))] = os.path.join(root, name)
    return found


def read_workbook(path: str, reporting_year: int) -> list:
    """Read the baseball-relevant columns out of one EADA workbook."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise ImportError(
            "Reading EADA workbooks needs openpyxl: "
            'pip install "ncaa_bbStats[scrape]"'
        ) from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    index = {name: header.index(name) for name in SOURCE_COLUMNS if name in header}
    missing = [name for name in SOURCE_COLUMNS if name not in index]
    if missing:
        raise SystemExit(f"{os.path.basename(path)} is missing columns: {missing}")

    out = []
    for row in rows:
        record = {name: row[position] for name, position in index.items()}
        # Only institutions that actually field a baseball team.
        participants = _number(record.get("PARTIC_MEN_Baseball"))
        if not participants:
            continue
        record["eada_year"] = reporting_year
        out.append(record)
    workbook.close()
    return out


def derive_features(records: list) -> list:
    """Turn raw EADA columns into the twelve derived features."""
    expense = [_number(r.get("EXP_MEN_Baseball")) for r in records]
    revenue = [_number(r.get("REV_MEN_Baseball")) for r in records]
    opex = [_number(r.get("OPEXPPERPART_MEN_Baseball")) for r in records]
    recruiting = [_number(r.get("RECRUITEXP_MEN")) for r in records]
    salary = [_number(r.get("HDCOACH_SAL_FTE_MEN")) for r in records]

    roster = []
    for r in records:
        value = _number(r.get("PARTIC_MEN_Baseball"))
        roster.append(
            value if value is not None
            and PLAUSIBLE_ROSTER[0] <= value <= PLAUSIBLE_ROSTER[1]
            else None
        )

    budget_pct = _percentile_ranks(expense)
    opex_pct = _percentile_ranks(opex)
    recruiting_pct = _percentile_ranks(recruiting)

    out = []
    for i, record in enumerate(records):
        head = _number(record.get("MEN_TOTAL_HEADCOACH_Baseball")) or 0.0
        assistant = _number(record.get("MEN_TOTAL_ASSTCOACH_Baseball")) or 0.0
        per_player = (
            expense[i] / roster[i]
            if expense[i] is not None and roster[i] else None
        )
        out.append({
            "unitid": str(int(_number(record["unitid"]) or 0)),
            "institution_name": (record.get("institution_name") or "").strip(),
            "state": (record.get("state_cd") or "").strip(),
            "eada_year": record["eada_year"],
            "budget_pct": budget_pct[i],
            "log_budget": _log1p_nonneg(expense[i]),
            "opex_per_player_pct": opex_pct[i],
            "log_opex_per_player": _log1p_nonneg(opex[i]),
            "log_budget_per_player": _log1p_nonneg(per_player),
            "roster_size": roster[i],
            "log_revenue": _log1p_nonneg(revenue[i]),
            "net_revenue": (
                revenue[i] - expense[i]
                if revenue[i] is not None and expense[i] is not None else None
            ),
            "coaching_staff_size": head + assistant,
            "dept_recruiting_pct": recruiting_pct[i],
            "log_dept_recruiting": _log1p_nonneg(recruiting[i]),
            "log_dept_coach_salary": _log1p_nonneg(salary[i]),
        })
    return out


def attach_team_ids(features: list) -> tuple:
    """Join derived rows onto the registry by IPEDS unitid, then by name."""
    by_unitid = {}
    for row in _registry_rows():
        if row["ipeds_unitid"]:
            by_unitid[row["ipeds_unitid"]] = row["team_id"]

    matched, unmatched = [], []
    for row in features:
        team_id = by_unitid.get(row["unitid"])
        if team_id is None:
            team_id = resolve_team(row["institution_name"])
        if team_id is None:
            unmatched.append(row["institution_name"])
            continue
        row["team_id"] = team_id
        matched.append(row)
    return matched, unmatched


def _registry_rows():
    from ncaa_bbStats.team_registry import _teams
    return list(_teams().values())


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--eada-dir", required=True,
                        help='Directory containing EADA_<YYYY>.xlsx files.')
    parser.add_argument("--out", default=data_path("program_finance"))
    args = parser.parse_args(argv)

    workbooks = find_workbooks(args.eada_dir)
    if not workbooks:
        raise SystemExit(
            f"No EADA_<YYYY>.xlsx found under {args.eada_dir}.\n"
            "Download the survey data from https://ope.ed.gov/athletics/."
        )

    all_rows = []
    for reporting_year in sorted(workbooks):
        if reporting_year not in REPORTING_YEARS:
            continue
        print(f"reading EADA_{reporting_year}.xlsx ...")
        records = read_workbook(workbooks[reporting_year], reporting_year)
        features = derive_features(records)
        matched, unmatched = attach_team_ids(features)
        print(f"  {len(records):5d} institutions with baseball, "
              f"{len(matched):4d} matched to programs "
              f"({len(unmatched)} unmatched)")

        for row in matched:
            all_rows.append(dict(row, season=reporting_year,
                                 carried_forward=False))

    # Carry the most recent survey forward into seasons it does not cover.
    by_year = {}
    for row in all_rows:
        by_year.setdefault(row["season"], []).append(row)
    for target, source in CARRY_FORWARD.items():
        for row in by_year.get(source, []):
            all_rows.append(dict(row, season=target, carried_forward=True))
        print(f"carried {source} forward to {target} "
              f"({len(by_year.get(source, []))} programs)")

    os.makedirs(args.out, exist_ok=True)
    path = os.path.join(args.out, "eada_features.csv")
    all_rows.sort(key=lambda r: (r["season"], r["team_id"]))
    with open(path, "w", newline="\n", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS,
                                extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\nwrote {len(all_rows)} team-seasons -> {os.path.relpath(path)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
